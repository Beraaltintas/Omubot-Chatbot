import warnings
warnings.filterwarnings('ignore')
import json
import pandas as pd
import random
import torch
from torch.utils.data import Dataset, DataLoader
from transformers import BertTokenizer, BertForSequenceClassification, Trainer, TrainingArguments
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, precision_recall_fscore_support
from nltk import word_tokenize
from nltk.stem import PorterStemmer
import nltk
from openpyxl import load_workbook

# NLTK veri indirme
nltk.download('punkt')

def load_json_file(filename):
    with open(filename, encoding='utf-8') as f:
        file = json.load(f)
    return file

filename = 'C:\\Users\\mbera\\Desktop\\omubot\\verisetison1.json'  # Dosya yolunu güncelleyin
intents = load_json_file(filename)

def create_df():
    return pd.DataFrame(columns=['Pattern', 'Tag', 'Response'])

df = create_df()

def extract_json_info(json_file, df):
    new_rows = []
    for intent in json_file['intents']:
        if 'patterns' in intent:  # patterns anahtarının varlığını kontrol et
            for pattern in intent['patterns']:
                for response in intent['responses']:
                    new_rows.append({'Pattern': pattern, 'Tag': intent['tag'], 'Response': response})
    
    new_df = pd.DataFrame(new_rows)
    df = pd.concat([df, new_df], ignore_index=True)
    return df


df = extract_json_info(intents, df)
df.head()

df2 = df.copy()

stemmer = PorterStemmer()
ignore_words = ['?', '!', ',', '.']

def preprocess_pattern(pattern):
    words = word_tokenize(pattern.lower())
    stemmed_words = [stemmer.stem(word) for word in words if word not in ignore_words]
    return " ".join(stemmed_words)

df['Pattern'] = df['Pattern'].apply(preprocess_pattern)

labels = df2['Tag'].unique().tolist()
labels = [s.strip() for s in labels]
num_labels = len(labels)
id2label = {id: label for id, label in enumerate(labels)}
label2id = {label: id for id, label in enumerate(labels)}

df2['labels'] = df2['Tag'].map(lambda x: label2id[x.strip()])

X = list(df2['Pattern'])
y = list(df2['labels'])

X_train, X_test, y_train, y_test = train_test_split(X, y, random_state=123)

model_name = "dbmdz/bert-base-turkish-cased"
max_len = 256

tokenizer = BertTokenizer.from_pretrained(model_name, max_length=max_len)
model = BertForSequenceClassification.from_pretrained(model_name, num_labels=num_labels, id2label=id2label, label2id=label2id)

train_encoding = tokenizer(X_train, truncation=True, padding=True, max_length=max_len)
test_encoding = tokenizer(X_test, truncation=True, padding=True, max_length=max_len)

class ChatDataset(Dataset):
    def __init__(self, encodings, labels):
        self.encodings = encodings
        self.labels = labels

    def __getitem__(self, idx):
        item = {key: torch.tensor(val[idx]) for key, val in self.encodings.items()}
        item['labels'] = torch.tensor(self.labels[idx])
        return item

    def __len__(self):
        return len(self.labels)

train_dataset = ChatDataset(train_encoding, y_train)
test_dataset = ChatDataset(test_encoding, y_test)

def compute_metrics(pred):
    labels = pred.label_ids
    preds = pred.predictions.argmax(-1)
    precision, recall, f1, _ = precision_recall_fscore_support(labels, preds, average='macro')
    acc = accuracy_score(labels, preds)
    return {
        'Accuracy': acc,
        'F1': f1,
        'Precision': precision,
        'Recall': recall
    }

training_args = TrainingArguments(
    output_dir='./output',
    do_train=True,
    do_eval=True,
    num_train_epochs=25,
    per_device_train_batch_size=16,
    per_device_eval_batch_size=8,
    warmup_steps=100,
    weight_decay=0.05,
    logging_strategy='steps',
    logging_dir='./logs',
    logging_steps=50,
    evaluation_strategy="steps",
    eval_steps=50,
    save_strategy="steps",
    load_best_model_at_end=True
)


trainer = Trainer(
    model=model,
    args=training_args,
    train_dataset=train_dataset,
    eval_dataset=test_dataset,
    compute_metrics=compute_metrics
)

trainer.train()

eval_results = trainer.evaluate()
print(f"Değerlendirme Sonuçları: {eval_results}")

def log_unknown_question(question, file_path='C:\\Users\\mbera\\Desktop\\omubot\\unknown_questions.xlsx'):
    try:
        # Excel dosyasını yükle
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        writer.book = book
        
        # Soruyu DataFrame olarak oluştur
        df = pd.DataFrame([[question]], columns=['Question'])
        
        # Yeni veriyi mevcut dosyaya ekle
        df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row, sheet_name='Sheet1')
        
        # Dosyayı kaydet ve kapat
        writer.save()
        writer.close()
    except FileNotFoundError:
        # Eğer dosya yoksa yeni bir dosya oluştur
        df = pd.DataFrame([[question]], columns=['Question'])
        df.to_excel(file_path, index=False, sheet_name='Sheet1')

def predict(text):
    # Modeli kullanarak tahminde bulun
    inputs = tokenizer(text, padding=True, truncation=True, max_length=512, return_tensors="pt")
    outputs = model(**inputs)
    probs = outputs.logits.softmax(dim=1)
    pred_label_idx = probs.argmax()
    pred_label = model.config.id2label[pred_label_idx.item()]

    # Etiketlerin DataFrame'de olup olmadığını kontrol et
    if pred_label not in df['Tag'].values:
        log_unknown_question(text)  # Bilinmeyen soruyu kaydet
        return "Bu sorunuzu yanıtlayamıyorum, en kısa sürede yanıtlayacağım."

    # Yanıtları DataFrame'den çek
    responses = df[df['Tag'] == pred_label]['Response']

    # Yanıtların boş olup olmadığını kontrol et
    if responses.empty:
        log_unknown_question(text)  # Bilinmeyen soruyu kaydet
        return "Bu sorunuzu yanıtlayamıyorum, en kısa sürede yanıtlayacağım."

    # Yanıtları doğru şekilde seç ve UTF-8'e dönüştür
    response = random.choice(responses.tolist())

    # UTF-8 encoding ve decoding işlemleri
    response = response.encode('utf-8').decode('utf-8', errors='ignore')

    return response

model_path = "C:\\Users\\mbera\\Desktop\\omubot\\omubotonlineSon1"
trainer.save_model(model_path)
tokenizer.save_pretrained(model_path)

if __name__ == "__main__":
    while True:
        text = input("Soru: ")
        if text.lower() in ['exit', 'quit', 'q']:
            break
        response = predict(text)
        print(f"Cevap: {response}")
